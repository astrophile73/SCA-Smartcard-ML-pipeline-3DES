import re
from typing import Dict, Optional, Tuple


def _norm_hex(value: str) -> str:
    v = str(value or "").strip().replace(" ", "").upper()
    return "".join(ch for ch in v if ch in "0123456789ABCDEF")


def _track2_pan(track2: str) -> str:
    t2 = _norm_hex(track2)
    for sep in ("D", "F", "="):
        if sep in t2:
            return t2.split(sep, 1)[0]
    return t2


def _profile_from_text(text: str) -> Optional[str]:
    v = str(text or "").strip().lower()
    if "visa" in v:
        return "visa"
    if "master" in v:
        return "mastercard"
    return None


def _valid_key(k: str) -> bool:
    v = _norm_hex(k)
    return len(v) == 32 and not re.fullmatch(r"0+", v)


def load_external_3des_label_map(xlsx_path: str) -> Dict[str, Dict[str, str]]:
    """
    Load PROFILE/TRACK2 -> 3DES_KENC/KMAC/KDEK mapping from XLSX.
    Returns:
      {
        "<profile>|<track2_full>": {"T_DES_KENC": "...", ...},
        "<profile>|PAN:<pan>": {...}
      }
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if not wb.sheetnames:
        return {}

    ws = wb[wb.sheetnames[0]]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    required = ["PROFILE", "TRACK2", "3DES_KENC", "3DES_KMAC", "3DES_KDEK"]
    for col in required:
        if col not in idx:
            raise ValueError(f"Missing required column '{col}' in label map: {xlsx_path}")

    out: Dict[str, Dict[str, str]] = {}
    per_profile = {"visa": set(), "mastercard": set()}
    for r in range(2, ws.max_row + 1):
        profile = _profile_from_text(ws.cell(r, idx["PROFILE"]).value)
        track2 = _norm_hex(str(ws.cell(r, idx["TRACK2"]).value or ""))
        if not profile or not track2:
            continue

        kenc = _norm_hex(str(ws.cell(r, idx["3DES_KENC"]).value or ""))
        kmac = _norm_hex(str(ws.cell(r, idx["3DES_KMAC"]).value or ""))
        kdek = _norm_hex(str(ws.cell(r, idx["3DES_KDEK"]).value or ""))
        if not (_valid_key(kenc) and _valid_key(kmac) and _valid_key(kdek)):
            continue

        row_keys = {"T_DES_KENC": kenc, "T_DES_KMAC": kmac, "T_DES_KDEK": kdek}
        out[f"{profile}|{track2}"] = row_keys
        pan = _track2_pan(track2)
        if pan:
            out[f"{profile}|PAN:{pan}"] = row_keys
        per_profile[profile].add((kenc, kmac, kdek))

    for profile, combos in per_profile.items():
        if len(combos) == 1:
            kenc, kmac, kdek = next(iter(combos))
            out[f"{profile}|ANY"] = {"T_DES_KENC": kenc, "T_DES_KMAC": kmac, "T_DES_KDEK": kdek}

    return out


def resolve_external_3des_keys(
    label_map: Dict[str, Dict[str, str]],
    profile: Optional[str],
    track2: str,
) -> Optional[Tuple[str, str, str]]:
    if not label_map:
        return None
    p = _profile_from_text(profile or "")
    if not p:
        return None

    t2 = _norm_hex(track2)
    pan = _track2_pan(t2)
    for key in (f"{p}|{t2}", f"{p}|PAN:{pan}", f"{p}|ANY"):
        row = label_map.get(key)
        if row:
            return row["T_DES_KENC"], row["T_DES_KMAC"], row["T_DES_KDEK"]
    return None
